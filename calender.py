# Import necessary libraries for GUI and Excel handling
from customtkinter import *  # CustomTkinter for enhanced tkinter widgets
from tkinter import messagebox, filedialog, simpledialog  # Standard tkinter dialogs
from tkcalendar import DateEntry  # Calendar widget for date selection
from datetime import datetime, timedelta  # For date manipulation
import openpyxl  # Library for handling Excel files
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Styles for Excel cells
from openpyxl.utils import get_column_letter  # Utility to convert column numbers to letters

# Global variables to store event data and UI components
date_label_dict = {}  # Dictionary to map dates to their corresponding labels
added_events = []  # List to keep track of added events
events_window = None  # Window for displaying all events
events_textbox = None  # Textbox to show events in the events window

def show_events_window():
    """Creates and displays a window summarizing the number of working days by weekdays."""
    global events_window, events_textbox
    if events_window is None or not events_window.winfo_exists():
        events_window = CTkToplevel()
        events_window.title("Working Days Summary")
        events_window.geometry("400x400")

        title_label = CTkLabel(events_window, text="Working Days Summary", font=("Arial", 18, "bold"))
        title_label.pack(pady=10)

        events_textbox = CTkTextbox(events_window, width=380, height=350)
        events_textbox.pack(padx=10, pady=5)

    # Update the textbox content
    if events_textbox:
        events_textbox.configure(state="normal")
        events_textbox.delete("1.0", END)

        # Display working days summary
        working_days = get_working_days_by_weekday()
        summary_text = "Working Days by Weekday:\n\n"
        for day, data in working_days.items():
            summary_text += f"{day} ({data['count']}):\n"
            for date in data["dates"]:
                summary_text += f"  - {date}\n"
            summary_text += "\n"

        events_textbox.insert(END, summary_text)
        events_textbox.configure(state="disabled")

def update_working_days_display():
    """Updates the Working Days textbox with the summary of working days."""
    working_days_textbox.configure(state="normal")
    working_days_textbox.delete("1.0", END)

    working_days = get_working_days_by_weekday()
    summary_text = "Working Days by Weekday:\n\n"
    for day, data in working_days.items():
        summary_text += f"{day} ({data['count']}):\n"
        for date in data["dates"]:
            summary_text += f"  - {date}\n"
        summary_text += "\n"

    working_days_textbox.insert(END, summary_text)
    working_days_textbox.configure(state="disabled")



def update_events_display():
    """Updates the events display in the events window."""
    if events_textbox and events_window and events_window.winfo_exists():
        events_textbox.configure(state="normal")  # Enable editing
        events_textbox.delete("1.0", END)  # Clear existing text
        
        events_by_date = {}  # Dictionary to group events by date
        # Populate events_by_date with events from date_label_dict
        for date_str, (_, label) in date_label_dict.items():
            event_text = label.cget("text")  # Get the text of the event label
            if event_text:
                if date_str not in events_by_date:
                    events_by_date[date_str] = []  # Initialize list for new date
                events_by_date[date_str].append(event_text)  # Append event text
        
        # Sort dates and display events
        for date_str in sorted(events_by_date.keys(), key=lambda x: datetime.strptime(x, "%d/%m/%y")):
            events = events_by_date[date_str]
            events_textbox.insert(END, f"\nDate: {date_str}\n")  # Insert date header
            events_textbox.insert(END, "─" * 30 + "\n")  # Insert separator
            for event in events:
                events_textbox.insert(END, f"• {event}\n")  # Insert each event
        
        events_textbox.configure(state="disabled")  # Disable editing

def update_frame():
    """Updates the calendar frame based on the selected date range."""
    global date_label_dict
    start_date = startDate.get_date()  # Get start date from DateEntry
    end_date = endDate.get_date()  # Get end date from DateEntry

    # Clear existing widgets in the date frame
    for widget in date_frame.winfo_children():
        widget.destroy()

    headers = ["Date", "Day", "Schedule"]  # Column headers for the calendar
    # Create header labels
    for col, header in enumerate(headers):
        CTkLabel(date_frame, text=header, font=("Consolas", 15, "bold")).grid(row=0, column=col, padx=5, sticky='w')

    # Separator line
    CTkLabel(date_frame, text="-" * 50, font=("Consolas", 15)).grid(row=1, column=0, columnspan=3, sticky='ew', pady=5)

    current_date = start_date  # Initialize current date
    row = 2  # Start from the third row in the grid
    date_label_dict = {}  # Reset date_label_dict for new date range

    # Loop through each date in the selected range
    while current_date <= end_date:
        date_str = current_date.strftime('%d/%m/%y')  # Format date as string
        day_str = current_date.strftime('%a')  # Get day of the week

        # Create labels for date, day, and schedule
        date_label = CTkLabel(date_frame, text=date_str, font=("Consolas", 15))
        day_label = CTkLabel(date_frame, text=day_str, font=("Consolas", 15))
        schedule_label = CTkLabel(date_frame, text="", font=("Consolas", 15))

        # Place labels in the grid
        date_label.grid(row=row, column=0, padx=5, sticky='w')
        day_label.grid(row=row, column=1, padx=5, sticky='w')
        schedule_label.grid(row=row, column=2, padx=5, sticky='w')

        # Determine schedule text based on the day of the week
        if current_date.weekday() == 6:  # Sunday
            schedule_label.configure(text="Sunday")
        elif current_date.weekday() == 5:  # Saturday
            saturday_number = (current_date.day - 1) // 7 + 1  # Calculate Saturday number
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(saturday_number % 10 * (saturday_number % 100 not in [11, 12, 13]), 'th')
            schedule_text = f"{saturday_number}{suffix} Saturday"  # Format Saturday text
            if saturday_number % 2 == 0:  # Even Saturday
                schedule_text += " : Holiday"
            schedule_label.configure(text=schedule_text)

        # Store the label in the date_label_dict
        date_label_dict[date_str] = (row, schedule_label)
        current_date += timedelta(days=1)  # Move to the next day
        row += 1  # Move to the next row

    # Call the function to display working days
    display_working_days()
    

def get_working_days_by_weekday():
    """Returns a dictionary of working days grouped by weekdays (Monday to Friday and Saturday) with counts."""
    working_days_by_weekday = {day: {"dates": [], "count": 0} for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]}

    # Iterate over date_label_dict to filter and categorize working days
    for date_str, (_, schedule_label) in date_label_dict.items():
        current_date = datetime.strptime(date_str, "%d/%m/%y")
        day_name = current_date.strftime("%A")  # Full name of the day (e.g., Monday)

        schedule_text = schedule_label.cget("text")
        if day_name == "Saturday":
            # Include Saturday only if it's not marked as a holiday
            if "Holiday" not in schedule_text:
                working_days_by_weekday[day_name]["dates"].append(date_str)
                working_days_by_weekday[day_name]["count"] += 1
        elif day_name not in ["Sunday"] and "Holiday" not in schedule_text:
            # Include weekdays (Monday to Friday) if they are not marked as holidays
            working_days_by_weekday[day_name]["dates"].append(date_str)
            working_days_by_weekday[day_name]["count"] += 1

    return working_days_by_weekday



def display_working_days():
    """Displays the working days grouped by weekdays with their counts, dynamically handling Saturdays."""
    working_days = get_working_days_by_weekday()
    working_days_text = "Working Days by Weekday:\n\n"
    
    for day, data in working_days.items():
        working_days_text += f"{day} ({data['count']}):\n"
        for date in data["dates"]:
            working_days_text += f"  - {date}\n"
        working_days_text += "\n"  # Add a blank line for separation

    # Display in the global_textbox
    global_textbox.configure(state="normal")
    global_textbox.delete("1.0", END)
    global_textbox.insert(END, working_days_text)
    global_textbox.configure(state="disabled")


def add_event_to_calendar(row, event):
    """Adds an event to the calendar on the specified row."""
    date_str = list(date_label_dict.keys())[row - 2]  # Get the date string from the dictionary
    schedule_label = date_label_dict[date_str][1]  # Get the corresponding schedule label
    current_text = schedule_label.cget("text")  # Get current text of the schedule label

    date = datetime.strptime(date_str, "%d/%m/%y")  # Convert date string to datetime object
    is_weekend = date.weekday() in [5, 6]  # Check if the date is a weekend (Saturday or Sunday)
    day_name = date.strftime("%A")  # Get the name of the day

    # Check if the event is a holiday
    is_holiday = "Holiday" in event

    if is_holiday:
        # If it's a holiday, set the schedule label as the holiday event
        schedule_label.configure(text=event)
    else:
        # Allow multiple events by appending them
        new_text = current_text + " | " + event if current_text else event

        # Split text into lines if it exceeds 50 characters
        if len(new_text) > 50:
            new_text = '\n'.join(new_text[i:i + 50] for i in range(0, len(new_text), 50))

        schedule_label.configure(text=new_text)  # Update the schedule label with the new text

    # Handle Saturday-specific logic
    if day_name == "Saturday":
        if is_holiday:
            # If a Saturday is marked as a holiday, update its label and remove it from working days
            schedule_label.configure(text=f"{event}")
        else:
            # Add a custom label for Saturdays if no holiday
            saturday_number = (date.day - 1) // 7 + 1  # Calculate which Saturday of the month
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(saturday_number, 'th')
            new_label = f"{saturday_number}{suffix} Saturday"
            schedule_label.configure(text=f"{new_label} | {event}" if event else new_label)

    # Store the updated schedule label in the date_label_dict
    date_label_dict[date_str] = (row, schedule_label)

    # Update the working days display after the event is added
    display_working_days()  # Refresh the working days display
    update_events_display()  # Refresh the events display in the UI

def selection(value):
    """Handles the selection of an event type and updates the UI accordingly."""
    for widget in frame4.winfo_children():
        widget.destroy()  # Clear previous widgets in the frame

    def on_add():
        """Handles the addition of an event when the 'add' button is clicked."""
        event_value = custom_event_entry.get() if value == "Others" else value  # Get event name
        if value == "Others" and not event_value:  # Check for empty custom event name
            messagebox.showerror("Error", "Please enter a custom event name.")
            return

        # Append ' (Holiday)' if the checkbox is selected
        if holiday_checkbox.get():
            event_value += " (Holiday)"

        # Display selected date or date range in the textbox
        if day_type.get() == "Single Day":
            selected_date = date_value.get_date().strftime("%d/%m/%y")
            textbox.configure(state="normal")
            textbox.delete("1.0", END)
            textbox.insert(END, f"Option: {event_value}\nDate: {selected_date}")
        else:
            start_date = start_date_value.get_date().strftime("%d/%m/%y")
            end_date = end_date_value.get_date().strftime("%d/%m/%y")
            textbox.configure(state="normal")
            textbox.delete("1.0", END)
            textbox.insert(END, f"Option: {event_value}\nStart Date: {start_date}\nEnd Date: {end_date}")
        textbox.configure(state="disabled")

    def on_submit():
        """Handles the submission of an event when the 'submit' button is clicked."""
        event_value = custom_event_entry.get() if value == "Others" else value  # Get event name
        if value == "Others" and not event_value:  # Check for empty custom event name
            messagebox.showerror("Error", "Please enter a custom event name.")
            return

        # Append ' (Holiday)' if the checkbox is selected
        if holiday_checkbox.get():
            event_value += " (Holiday)"

        # Add event to the calendar for a single day
        if day_type.get() == "Single Day":
            formatted_date = date_value.get_date().strftime("%d/%m/%y")
            if formatted_date in date_label_dict:
                row = list(date_label_dict.keys()).index(formatted_date) + 2  # Get row index
                add_event_to_calendar(row, event_value)  # Add event
            else:
                messagebox.showerror("Error", "Invalid date selected")
        else:  # Handle multiple days
            start_date = start_date_value.get_date()
            end_date = end_date_value.get_date()
            if start_date > end_date:  # Check for valid date range
                messagebox.showerror("Error", "Start date cannot be after end date.")
                return
            current_date = start_date
            while current_date <= end_date:  # Loop through each date in the range
                formatted_date = current_date.strftime("%d/%m/%y")
                if formatted_date in date_label_dict:
                    row = list(date_label_dict.keys()).index(formatted_date) + 2  # Get row index
                    add_event_to_calendar(row, event_value)  # Add event
                current_date += timedelta(days=1)  # Move to the next day

    # Display selected option and checkbox in the frame
    CTkLabel(frame4, text=f'Option selected: {value}').grid(row=0, column=0, sticky='nw', columnspan=2)

    if value == "Others":  # If 'Others' is selected, show custom event entry
        CTkLabel(frame4, text="Enter custom event name:").grid(row=1, column=0, sticky='w', pady=(10, 0))
        custom_event_entry = CTkEntry(frame4, width=300, placeholder_text="Type your event here")
        custom_event_entry.grid(row=2, column=0, sticky='w', padx=(250, 5), pady=(5, 10))

    # Checkbox for marking the event as a holiday
    holiday_checkbox = CTkCheckBox(frame4, text="Mark as Holiday")
    holiday_checkbox.grid(row=3, column=0, sticky='w', pady=(10, 0))

    # Dropdown for selecting single or multiple days
    day_type = CTkOptionMenu(frame4, values=["Single Day", "Multiple Days"], command=lambda x: update_date_fields())
    day_type.grid(row=2, column=0, sticky='nw', columnspan=2, pady=5)

    def update_date_fields():
        """Updates the date fields based on the selected day type."""
        for widget in frame4.winfo_children()[4:]:
            widget.destroy()  # Clear previous date fields

        if day_type.get() == "Single Day":  # Show single date selection
            CTkLabel(frame4, text='Select Date:').grid(row=4, column=0, sticky='nw', columnspan=2)
            global date_value
            date_value = DateEntry(frame4, date_pattern='dd/mm/yyyy')  # DateEntry for single date
            date_value.grid(row=4, column=0, sticky='w', padx=140, columnspan=2)
        else:  # Show start and end date selection
            CTkLabel(frame4, text='Start Date:').grid(row=4, column=0, sticky='nw', columnspan=2)
            global start_date_value
            start_date_value = DateEntry(frame4, date_pattern='dd/mm/yyyy')  # DateEntry for start date
            start_date_value.grid(row=4, column=0, sticky='w', padx=140, columnspan=2)

            CTkLabel(frame4, text='End Date:').grid(row=5, column=0, sticky='nw', columnspan=2)
            global end_date_value
            end_date_value = DateEntry(frame4, date_pattern='dd/mm/yyyy')  # DateEntry for end date
            end_date_value.grid(row=5, column=0, sticky='w', padx=140, columnspan=2)

        # Button to add the event
        CTkButton(frame4, text='add', corner_radius=1, height=15, width=60, command=on_add).grid(row=6, column=0, sticky='w', pady=5)

        global textbox
        textbox = CTkTextbox(frame4, state="disabled")  # Textbox to display selected event details
        textbox.grid(row=7, column=0, sticky='we', columnspan=2)
        frame4.grid_columnconfigure(0, weight=1)

        # Buttons for canceling and submitting the event
        CTkButton(frame4, text="Cancel", corner_radius=2, command=lambda: textbox.delete("1.0", END)).grid(row=8, column=0, sticky='w', pady=(16, 0))
        CTkButton(frame4, text="Submit", corner_radius=2, command=on_submit).grid(row=8, column=0, sticky='w', pady=(16, 0), padx=170)

    update_date_fields()  # Initialize date fields based on selection

def clear_last_event():
    """Clears the last added event from the calendar and updates the display."""
    if added_events:
        row, label = added_events.pop()  # Remove the last event
        label.configure(text="")  # Clear the label text
        global_textbox.configure(state="normal")
        lines = global_textbox.get("1.0", END).split('\n')  # Get all lines from the textbox
        global_textbox.delete("1.0", END)  # Clear the textbox
        for line in lines[:-2]:  # Reinsert all lines except the last two (the last event)
            global_textbox.insert(END, line + '\n')
        global_textbox.configure(state="disabled")  # Disable editing
        update_events_display()  # Refresh the events display
    else:
        messagebox.showinfo("Info", "No events to clear.")  # Inform if no events are present

def clear_calendar():
    """Clears the entire calendar and resets all data."""
    global date_label_dict, added_events
    startDate.set_date(datetime.now())  # Reset start date to today
    endDate.set_date(datetime.now())  # Reset end date to today
    added_events = []  # Clear added events
    for widget in date_frame.winfo_children():
        widget.destroy()  # Remove all widgets from the date frame
    date_label_dict = {}  # Reset date_label_dict
    global_textbox.configure(state="normal")
    global_textbox.delete("1.0", END)  # Clear the global textbox
    global_textbox.configure(state="disabled")
    update_events_display()  # Refresh the events display
    messagebox.showinfo("Info", "Calendar cleared successfully.")  # Inform user of success
def update_selected_events_display():
    """Updates the Selected Events textbox with the latest events."""
    global_textbox.configure(state="normal")
    global_textbox.delete("1.0", END)

    for date_str, (_, label) in date_label_dict.items():
        event_text = label.cget("text")
        if event_text:
            global_textbox.insert(END, f"{date_str}: {event_text}\n")

    global_textbox.configure(state="disabled")

def add_event_to_calendar(row, event):
    """Adds an event to the calendar and updates the displays."""
    date_str = list(date_label_dict.keys())[row - 2]  # Get the date string
    schedule_label = date_label_dict[date_str][1]  # Get the corresponding schedule label
    current_text = schedule_label.cget("text")

    # Update the schedule label
    new_text = current_text + " | " + event if current_text else event
    schedule_label.configure(text=new_text)

    # Refresh the displays
    update_selected_events_display()
    update_working_days_display()
    
def convert_to_excel():
    """Converts the current calendar to an Excel file."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Calendar"

    # Initialize row at the start
    row = 4  # Start from row 4 (after headers)

    # Define cell styles
    header_fill = PatternFill(start_color="4F6228", end_color="4F6228", fill_type="solid")  # Header fill color
    subheader_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Subheader fill color
    holiday_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Holiday fill color
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue fill for academic events
    light_violet_fill = PatternFill(start_color="E4D7F1", end_color="E4D7F1", fill_type="solid")  # Light violet fill for cultural events
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Cell border style


    
    # Define events that should be colored light blue (academic events)
    light_blue_events = [
        'Enrolment and commencement of classes for all UG and PG / commencement',
        'Finalisation of electives',
        'First Class committee meeting',
        'Commencement of Mid-Semester Exam',
        'Completion of quizzes, midsem and attendance entry in AUMS',
        'Second class committee',
        'Missed mid semester exam',
        'Pre-registration for next sem, course end survey, faculty feedback',
        'Finalisation of internals and attendance',
        'Last instruction day',
        'Commencement of end-semester exams'
    ]

    # Define events that should be colored light violet (festivals/cultural events)
    light_violet_events = [
        'Sree Krishna Janmashtami',
        'Ganesh Chaturthi',
        'Deepavali'
    ]

    # Define holiday events
    holiday_events = [
        'Sree Krishna Janmashtami',
        'Ganesh Chaturthi',
        'Deepavali'
    ]

    # Get semester dates
    sem_dates = {
        'UG-S1': (datetime.combine(sem1_start.get_date(), datetime.min.time()),
                  datetime.combine(sem1_end.get_date(), datetime.min.time())),
        'UG-S3': (datetime.combine(sem3_start.get_date(), datetime.min.time()),
                  datetime.combine(sem3_end.get_date(), datetime.min.time())),
        'UG-S5': (datetime.combine(sem5_start.get_date(), datetime.min.time()),
                  datetime.combine(sem5_end.get_date(), datetime.min.time())),
        'UG-S7': (datetime.combine(sem7_start.get_date(), datetime.min.time()),
                  datetime.combine(sem7_end.get_date(), datetime.min.time())),
        'PG-S1': (datetime.combine(pg_sem1_start.get_date(), datetime.min.time()),
                  datetime.combine(pg_sem1_end.get_date(), datetime.min.time())),
        'PG-S3': (datetime.combine(pg_sem3_start.get_date(), datetime.min.time()),
                  datetime.combine(pg_sem3_end.get_date(), datetime.min.time()))
    }

    # Group semesters by start date
    semesters_by_start_date = {}
    for sem_name, (start_date, end_date) in sem_dates.items():
        start_date_str = start_date.strftime("%Y-%m-%d")
        if start_date_str not in semesters_by_start_date:
            semesters_by_start_date[start_date_str] = {
                'semesters': [],
                'start_date': start_date,
                'end_date': end_date
            }
        semesters_by_start_date[start_date_str]['semesters'].append(sem_name)
        if end_date > semesters_by_start_date[start_date_str]['end_date']:
            semesters_by_start_date[start_date_str]['end_date'] = end_date

    # Initialize working day counters for each semester
    working_days = {
        'UG-S1': {'count': 0, 'start_date': None},
        'UG-S3': {'count': 0, 'start_date': None},
        'UG-S5': {'count': 0, 'start_date': None},
        'UG-S7': {'count': 0, 'start_date': None},
        'PG-S1': {'count': 0, 'start_date': None},
        'PG-S3': {'count': 0, 'start_date': None}
    }

    headers = ["Date", "Day", "UG-S1", "UG-S3", "UG-S5", "UG-S7", "PG-S1", "PG-S3", "Events/Holidays"]

    # Create merged headers
    merged_headers = ["Date", "Day"]
    merged_column_mapping = {}
    current_col = 3

    for start_date_str, group_info in sorted(semesters_by_start_date.items()):
        sems = group_info['semesters']
        if len(sems) > 1:
            header = " & ".join(sems)
            merged_headers.append(header)
            for sem in sems:
                merged_column_mapping[headers.index(sem) + 1] = current_col
        else:
            merged_headers.append(sems[0])
            merged_column_mapping[headers.index(sems[0]) + 1] = current_col
        current_col += 1

    merged_headers.append("Events/Holidays")

    total_columns = len(merged_headers)

    # Write merged headers
    for col, header in enumerate(merged_headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Merge cells for the title and subtitle based on the number of semesters
    for r in range(1, 3):
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_columns)  # Merge cells for title and subtitle
        cell = sheet.cell(row=r, column=1)  # Get the first column of the merged cells
        cell.font = Font(bold=True, size=14 if r == 1 else 12)
        cell.alignment = Alignment(horizontal='center')

    # Set title and subtitle
    sheet['A1'] = 'Amrita School of Engineering Bengaluru'
    sheet['A2'] = 'ACADEMIC CALENDAR (2024 - 2025) ODD SEMESTER'
    title_cell = sheet['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    subtitle_cell = sheet['A2']
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal='center')

    for i, width in enumerate([15] * (total_columns - 1) + [40], start=1):  # Last column (Events/Holidays) wider
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Initialize variables for event column properties
    event_column_index = total_columns  # Last column for events
    event_column_width = 40  # Width for the event column
    sheet.column_dimensions[get_column_letter(event_column_index)].width = event_column_width

    # Initialize tracking variables
    current_month = ""
    saturday_count_per_month = 0
    current_event = None
    event_start_row = None
    events_col = total_columns  # Last column for events

    # Process each date
    dates_list = sorted(date_label_dict.items(), key=lambda x: datetime.strptime(x[0], "%d/%m/%y"))
    
    for date_str, (_, schedule_label) in dates_list:
        date = datetime.strptime(date_str, "%d/%m/%y")
        event_text = schedule_label.cget("text")

        # Handle month headers
        if current_month != date.strftime("%B - %Y"):
            current_month = date.strftime("%B - %Y")
            # Create a new row for the month header
            row += 1
            # Merge cells across all columns for the month header
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
            month_cell = sheet.cell(row=row, column=1, value=current_month)
            month_cell.font = Font(bold=True)
            month_cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange color
            month_cell.alignment = Alignment(horizontal='center', vertical='center')
            month_cell.border = border  # Add border to the merged cell
            
            # Add borders to all cells in the merged range
            for col in range(1, total_columns + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
            
            row += 1
            saturday_count_per_month = 0

        # Write date and day
        date_cell = sheet.cell(row=row, column=1, value=date.strftime("%d-%b"))
        day_cell = sheet.cell(row=row, column=2, value=date.strftime("%a"))

        # Apply borders and alignment
        for col in range(1, total_columns + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Handle working days logic
        is_working_day = True

        # Handle Saturdays
        if date.weekday() == 5:  # Saturday
            saturday_count_per_month += 1
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(saturday_count_per_month, 'th')

            if saturday_count_per_month in [2, 4]:
                if event_text and not event_text.endswith("Saturday : Holiday"):
                    is_working_day = True
                    sheet.cell(row=row, column=events_col, value=event_text)
                else:
                    for col in range(1, total_columns + 1):
                        sheet.cell(row=row, column=col).fill = holiday_fill
                    is_working_day = False
                    sheet.cell(row=row, column=events_col, 
                             value=f"{saturday_count_per_month}{suffix} Saturday : Holiday")
            else:
                is_working_day = True
                if event_text and not event_text.endswith("Saturday"):
                    sheet.cell(row=row, column=events_col, value=event_text)
                else:
                    sheet.cell(row=row, column=events_col, 
                             value=f"{saturday_count_per_month}{suffix} Saturday")

                if event_text and any(holiday in event_text for holiday in holiday_events):
                    for col in range(1, total_columns + 1):
                        sheet.cell(row=row, column=col).fill = holiday_fill
                    is_working_day = False

        elif date.weekday() == 6:  # Sunday
            for col in range(1, total_columns + 1):
                sheet.cell(row=row, column=col).fill = holiday_fill
            is_working_day = False
            sheet.cell(row=row, column=events_col, value="Sunday")

        # Handle working day counts
        if is_working_day:
            for start_date_str, group_info in semesters_by_start_date.items():
                sems = group_info['semesters']
                start_date = group_info['start_date']
                end_date = group_info['end_date']
                
                if start_date <= date <= end_date:
                    merged_col = None
                    for sem in sems:
                        orig_col = headers.index(sem) + 1
                        merged_col = merged_column_mapping[orig_col]
                        working_days[sem]['count'] += 1
                        if working_days[sem]['start_date'] is None:
                            working_days[sem]['start_date'] = date
            
                    if merged_col is not None:
                        counts = [str(working_days[sem]['count']) for sem in sems]
                        if all(counts[0] == count for count in counts):
                            sheet.cell(row=row, column=merged_col, value=counts[0])
                        else:
                            sheet.cell(row=row, column=merged_col, value="/".join(counts))

        # Handle events and merging
        if event_text:
            event_cell = sheet.cell(row=row, column=events_col, value=event_text)
            event_cell.alignment = Alignment(horizontal='left', wrap_text=True)

            # Check if this is a continuation of the current event
            if event_text == current_event:
                # Don't set the value again, just track the row
                sheet.cell(row=row, column=events_col, value="")

                # Keep the event text even on weekends
                if date.weekday() in [5, 6]:  # Saturday or Sunday
                    sheet.cell(row=row, column=events_col, value="")
                
            else:
                # If there was a previous event, merge its cells
                if current_event is not None and event_start_row is not None:
                    sheet.merge_cells(start_row=event_start_row, start_column=events_col, 
                                    end_row=row-1, end_column=events_col)
                # Start tracking new event
                current_event = event_text
                event_start_row = row

            # Apply appropriate fills based on event type
            if any(event in event_text for event in light_blue_events):
                for col in range(1, total_columns + 1):
                    sheet.cell(row=row, column=col).fill = light_blue_fill
            elif any(event in event_text for event in light_violet_events):
                for col in range(1, total_columns + 1):
                    sheet.cell(row=row, column=col).fill = light_violet_fill
            elif any(event in event_text for event in holiday_events):
                for col in range(1, total_columns + 1):
                    sheet.cell(row=row, column=col).fill = holiday_fill
        else:
            # If there was an event and it's ending, merge its cells
            if current_event is not None and event_start_row is not None:
                sheet.merge_cells(start_row=event_start_row, start_column=events_col, 
                                end_row=row-1, end_column=events_col)
                current_event = None
                event_start_row = None

        row += 1

    # Merge cells for any final event that reaches the end
    if current_event is not None and event_start_row is not None:
        sheet.merge_cells(start_row=event_start_row, start_column=events_col, 
                         end_row=row-1, end_column=events_col)


# Add Working Days Breakdown by Weekday
    row += 2  # Leave some space after the last data row
    breakdown_title = sheet.cell(row=row, column=1, value="Working Days Breakdown by Weekday")
    breakdown_title.font = Font(bold=True, size=12)
    breakdown_title.alignment = Alignment(horizontal='center')
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
    for col in range(1, total_columns + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = border  # Add border to the title cell
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    row += 1

    # Retrieve working days by weekday
    working_days_by_weekday = get_working_days_by_weekday()

    for day, data in working_days_by_weekday.items():
        # Display the weekday and its count
        weekday_cell = sheet.cell(row=row, column=1, value=f"{day}: {data['count']} days")
        weekday_cell.font = Font(bold=True)
        weekday_cell.alignment = Alignment(horizontal='left')
        weekday_cell.border = border

        # List dates for each weekday
        dates_cell = sheet.cell(row=row, column=2, value=", ".join(data["dates"]))
        dates_cell.alignment = Alignment(horizontal='left', wrap_text=True)
        dates_cell.border = border
        row += 1

    # Add Total Working Days Below Each Column
    total_working_days_row = row + 2  # Leave some space after the last data row

    # Title for Total Working Days
    title_cell = sheet.cell(row=total_working_days_row, column=1, value="Total Working Days")
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=total_working_days_row, start_column=1, end_row=total_working_days_row, end_column=2)

    for sem_name, data in working_days.items():
        # Place total working days below the respective semester column
        column_index = headers.index(sem_name) + 1  # Find the correct column
        total_cell = sheet.cell(row=total_working_days_row, column=column_index, value=data["count"])
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.border = border



    # Save the Excel file
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        try:
            wb.save(file_path)
            messagebox.showinfo("Success", "Data successfully exported to Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file. Error: {e}")

# Main application setup
app = CTk()  # Create the main application window
app.title("Calendar Generator")  # Set window title
app.geometry("1620x650+0+0")  # Set window size
set_appearance_mode("dark")  # Set appearance mode to dark

#Scrollabe Frame
scrollable_frame = CTkScrollableFrame(app,height=650)  # Create a scrollable frame
scrollable_frame.grid(row=0, column=0, sticky='nsew')  # Place it in the grid
scrollable_frame.grid_columnconfigure(0, weight=1)  # Configure column weight
scrollable_frame.grid_rowconfigure(0, weight=1)

# Create main frame for the application
frame = CTkFrame(scrollable_frame, corner_radius=10)
frame.grid(row=0, column=0, sticky='nsew', padx=0, pady=0)
app.grid_columnconfigure(0, weight=1)  # Configure column weight
app.grid_rowconfigure(0, weight=1)  # Configure row weight

# Right side - Calendar view (remains unchanged)
date_frame_title = CTkLabel(app, text="Generated Calendar", font=("Arial", 18, "bold"))
date_frame_title.grid(row=0, column=1, sticky='n', padx=4, pady=(5, 0))
date_frame = CTkScrollableFrame(app, width=550, height=590, fg_color='#2b2b2b', corner_radius=10)
date_frame.grid(row=0, column=1, sticky='nse', padx=(5,15), pady=(35, 60))

frame1 = CTkFrame(app, width=150, height=40, fg_color='transparent')
frame1.grid(row=0, column=1, sticky='swe', pady=4, padx=4)
# Add close button to the frame
close_button = CTkButton(frame1, height=38, text='Close', corner_radius=5, command=app.destroy)
close_button.grid(row=0, column=3, padx=8, pady=(0,10))

# Left side - Modified layout
frame2 = CTkFrame(frame, height=120)
frame2.grid(row=0, column=0, padx=4, pady=4, sticky='nwe')

# Reduced height for event addition section
frame3 = CTkFrame(frame, height=150)  # Reduced height
frame3.grid(row=1, column=0, sticky='new', padx=4, pady=4)
frame.grid_columnconfigure(0, weight=1)
frame4 = CTkFrame(frame3, fg_color='transparent', height=120)  # Reduced height
frame4.grid(row=1, column=0, columnspan=2, padx=10, pady=4, sticky='we')
frame4.grid_columnconfigure(0, weight=1)

# Increased height for Selected Events section
global_textbox_frame = CTkFrame(frame)
global_textbox_title = CTkLabel(global_textbox_frame, text="Selected Events", font=("Arial", 16, "bold"))
global_textbox_title.pack(pady=(5, 0))

global_textbox_frame.grid(row=2, column=0, sticky='nsew', padx=5, pady=4)
frame.grid_rowconfigure(2, weight=3)  # Increased weight for more space


# Working Days Title and Textbox

working_days_textbox = CTkTextbox(global_textbox_frame, height=150, state="disabled")  # New textbox for working days
working_days_textbox.pack(expand=True, fill='both', padx=5, pady=5)
working_days_title = CTkLabel(global_textbox_frame, text="Working Days Summary", font=("Arial", 16, "bold"))
working_days_title.pack(pady=(10, 0))


global_textbox = CTkTextbox(global_textbox_frame, height=250, state="disabled")  # Increased height
global_textbox.pack(expand=True, fill='both', padx=5, pady=5)

# Rest of the UI elements
options = [
    'Enrolment and commencement of classes for all UG and PG / commencement',
    'Enrolment and commencement of classes for all UG-S1,PG-S1 / commencement ',
    'Finalisation of electives',
    'First Class committee meeting',
    'Commencement of Mid-Semester Exam',
    'Completion of quizzes, midsem and attendance entry in AUMS',
    'Second class committee',
    'Missed mid semester exam',
    'Pre-registration for next sem, course end survey, faculty feedback',
    'Finalisation of internals and attendance',
    'Last instruction day',
    'Commencement of end-semester exams',
    'Sree Krishna Janmashtami',
    'Ganesh Chaturthi',
    'Deepavali',
    'Others'
]

# Dropdown menu for selecting event type
option = CTkOptionMenu(frame3, values=options, corner_radius=1, command=selection)
option.grid(row=0, column=0, padx=10, pady=10, sticky='w')

# ... existing code ...

# Calendar range label and date entry fields
label = CTkLabel(frame2, text="Calendar Range", font=("Arial", 18))
label.grid(row=0, column=0, columnspan=2, sticky='w', padx=10, pady=10)

# Calendar range fields
sLabel = CTkLabel(frame2, text="Start Date:")
sLabel.grid(row=1, column=0, stick='w', padx=10, pady=0)
startDate = DateEntry(frame2, date_pattern="dd/mm/yyyy")
startDate.grid(row=1, column=1, sticky='w')
eLabel = CTkLabel(frame2, text="End Date:")
eLabel.grid(row=1, column=2, stick='w', padx=10, pady=0)
endDate = DateEntry(frame2, date_pattern="dd/mm/yyyy")
endDate.grid(row=1, column=3, sticky='w')

# Semester dates section
sem_label = CTkLabel(frame2, text="Semester Dates", font=("Arial", 16))
sem_label.grid(row=2, column=0, columnspan=2, sticky='w', padx=10, pady=(20,5))

# Semester 1
sem1_label = CTkLabel(frame2, text="Semester 1:")
sem1_label.grid(row=3, column=0, sticky='w', padx=10, pady=2)
sem1_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem1_start.grid(row=3, column=1, sticky='w')
sem1_end_label = CTkLabel(frame2, text="-----------------------")
sem1_end_label.grid(row=3, column=2, sticky='w', padx=5)
sem1_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem1_end.grid(row=3, column=3, sticky='w')

# Semester 3
sem3_label = CTkLabel(frame2, text="Semester 3:")
sem3_label.grid(row=4, column=0, sticky='w', padx=10, pady=2)
sem3_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem3_start.grid(row=4, column=1, sticky='w')
sem3_end_label = CTkLabel(frame2, text="-----------------------")
sem3_end_label.grid(row=4, column=2, sticky='w', padx=5)
sem3_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem3_end.grid(row=4, column=3, sticky='w')

# Semester 5
sem5_label = CTkLabel(frame2, text="Semester 5:")
sem5_label.grid(row=5, column=0, sticky='w', padx=10, pady=2)
sem5_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem5_start.grid(row=5, column=1, sticky='w')
sem5_end_label = CTkLabel(frame2, text="-----------------------")
sem5_end_label.grid(row=5, column=2, sticky='w', padx=5)
sem5_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem5_end.grid(row=5, column=3, sticky='w')

# Semester 7
sem7_label = CTkLabel(frame2, text="Semester 7:")
sem7_label.grid(row=6, column=0, sticky='w', padx=10, pady=2)
sem7_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem7_start.grid(row=6, column=1, sticky='w')
sem7_end_label = CTkLabel(frame2, text="-----------------------")
sem7_end_label.grid(row=6, column=2, sticky='w', padx=5)
sem7_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
sem7_end.grid(row=6, column=3, sticky='w')

# PG Semester dates section
pg_sem_label = CTkLabel(frame2, text="PG Semester Dates", font=("Arial", 16))
pg_sem_label.grid(row=7, column=0, columnspan=2, sticky='w', padx=10, pady=(20,5))

# PG Semester 1
pg_sem1_label = CTkLabel(frame2, text="PG Semester 1:")
pg_sem1_label.grid(row=8, column=0, sticky='w', padx=10, pady=2)
pg_sem1_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
pg_sem1_start.grid(row=8, column=1, sticky='w')
pg_sem1_end_label = CTkLabel(frame2, text="-----------------------")
pg_sem1_end_label.grid(row=8, column=2, sticky='w', padx=5)
pg_sem1_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
pg_sem1_end.grid(row=8, column=3, sticky='w')

# PG Semester 3
pg_sem3_label = CTkLabel(frame2, text="PG Semester 3:")
pg_sem3_label.grid(row=9, column=0, sticky='w', padx=10, pady=2)
pg_sem3_start = DateEntry(frame2, date_pattern="dd/mm/yyyy")
pg_sem3_start.grid(row=9, column=1, sticky='w')
pg_sem3_end_label = CTkLabel(frame2, text="-----------------------")
pg_sem3_end_label.grid(row=9, column=2, sticky='w', padx=5)
pg_sem3_end = DateEntry(frame2, date_pattern="dd/mm/yyyy")
pg_sem3_end.grid(row=9, column=3, sticky='w')

# Generate Calendar button (moved to bottom)
btn = CTkButton(frame2, text='Generate Calendar', corner_radius=4, height=30, width=320, command=update_frame)
btn.grid(row=10, column=0, columnspan=4, sticky='w', padx=10, pady=10)


# Buttons for clearing last event and clearing the calendar
clear = CTkButton(frame1, height=38, text='Clear Last Event', corner_radius=5, command=clear_last_event)
clear.grid(row=0, column=0, padx=2, pady=(0,10))
clear_cal = CTkButton(frame1, height=38, text='Clear Calendar', corner_radius=5, command=clear_calendar)
clear_cal.grid(row=0, column=1, padx=8, pady=(0,10))
excel = CTkButton(frame1, height=38, text='Convert to Excel', corner_radius=5, command=convert_to_excel)
excel.grid(row=0, column=2, pady=(0,10))

# Start the main application loop
app.mainloop()